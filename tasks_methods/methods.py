import csv
import json
import os
import re
import urllib.request
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from Log.logs import Logs
from openpyxl import Workbook
from robocorp import storage, workitems
from helpers.article import Article
from helpers.payload import Payload
from helpers.selector import Selector
from webdriver_util.webdrv_util import *
from selenium.webdriver.chrome.options import Options

load_dotenv("config\.env")
class ProducerMethods:
    @staticmethod
    def read_csv_create_work_item(debug:bool =True):
        csv_file_path = os.path.join("devdata", "csv_input.csv")

        if os.path.exists(csv_file_path):
            with open(csv_file_path, mode="r", newline="") as file:
                reader = csv.reader(file)
                header = next(reader)
                for row in reader:
                    payload = Payload(
                        phrase_test=row[0],
                        section=row[1],
                        data_range=int(row[2]),
                        sort_by=int(row[3]),
                        results=int(row[4])
                    )
                    if not debug:
                        workitems.outputs.create(
                            payload={
                                "phrase_test": payload.phrase_test,
                                "section": payload.section,
                                "data_range": payload.data_range,
                                "sort_by": payload.sort_by,
                                "results": payload.results
                            }
                        )
                    else:
                        return payload
        else:
            logger.critical(f"The CSV file: {csv_file_path} was not found.")


class ScraperMethods:
    @staticmethod
    def get_work_item() -> Payload | None:
        item = workitems.inputs.current
        if item:
            logger.info("Received payload:", item.payload)
            pay = Payload(
                phrase_test=item.payload["phrase_test"],
                section=item.payload["section"],
                data_range=int(item.payload["data_range"]),
                sort_by=int(item.payload["sort_by"]),
            )
            return pay
        else:
            logger.critical(f"Some error on process occurred!")

    @staticmethod
    def inicial_search(driver: Selenium, phrase: str):
        try:
            logger.info("Starting Scraper")
            search = find_element(
                        driver.driver, Selector(css='button[data-element="search-button"]')
                    )
            if search:
                center_element(driver.driver, search)
                click_elm(driver.driver, search)
                search_field = find_element(
                    driver.driver,
                    Selector(css="input[data-element='search-form-input']"),
                )
                if search_field:
                    center_element(driver.driver, search_field)
                    slow_send_keys(search_field, phrase + Keys.ENTER, False)
                    return True
        except Exception as e:
            print(e.with_traceback())
            return False

    @staticmethod
    def fine_search(
        driver: WebDriver,
        phrase: str,
        section: str,
        sort_by: int = 0,
        data_range: int = 0,
    ):
        try:
            no_results_match = find_element(
                driver.driver, 
                Selector(css="div[class='search-results-module-no-results']")
            )
            if no_results_match:
                logger.critical(f"Some error occurred: No search match...")
                return False
            #Expand Filter 
            label_search = find_element(
                driver.driver, Selector(css="span[class='see-all-text']")
            )
            if label_search:
                center_element(driver.driver, label_search)
                click_elm(driver.driver, label_search)
                if len(section.strip())>0:
                    list_topics = extract_names_from_list_items(driver)
                    if list_topics:
                        element_topic, topic = search_and_click_topics(driver.driver,list_topics, section)
                        if element_topic == False and topic == False:
                            return False           
                if sort_by>0: #not Relevance (default)
                    select_sort_by = find_element(
                        driver.driver, 
                        Selector(css="select[name='s']"))
                    if select_sort_by:
                        if sort_by == 1 or sort_by == 2:
                            select_option_value(
                                select_sort_by,
                                sort_by
                                )
                        else:
                            logger.error(f'Sort parameter not exists: {sort_by}')
                            logger.info('Relevance is selected')

                if data_range >= 0:
                    #!LEGEND: 0= Actual Page 1= Results you want 2= All Results 
                    if data_range == 0:
                        data_range_str = "Actual Page"
                    elif data_range == 1:
                        data_range_str = "Results you want"
                    elif data_range == 2:
                        data_range_str = "All Results"
                    if len(data_range_str.strip())>0:
                        logger.info(f'{data_range_str} is selected')
                    else:
                        data_range= 0                    
                        logger.info('Actual Page is selected')
            return True, data_range

        except Exception as e:
            logger.critical(f"Some error occurred: {e.__cause__}.")
            return False

    @staticmethod
    def collect_articles(driver: WebDriver, data_range:int=0) -> list[Article] | None:
        try:
            list_articles = []
            more_results = True
            while more_results:
                logger.info("Search results was found")
                search_results_section = find_element(
                    driver.driver,
                    Selector(css="div[class='search-results-module-results-header'"),
                )
                if search_results_section:
                    logger.info("Search results was found")
                    li_search_results = find_all_css(
                        driver.driver, 'ul[class*="search-results-module-results-menu"] li'
                    )
                    if li_search_results:
                        sleep(3)
                        lst_title = []
                        lt = TagAttVl(tag="a", attr="class", vlr="link") #title
                        lst_title.append(lt)
                        lst_description = []
                        ld = TagAttVl(tag="p", attr="class", vlr="promo-description")#description
                        lst_description.append(ld)
                        lst_time = []
                        ltm = TagAttVl(tag="p", attr="class", vlr="promo-timestamp") #time
                        lst_time.append(ltm)
                        for li in li_search_results:
                            logger.info("Create a article object")
                            article = Article()
                            title = find_elm_with_attribute(li, lst_title)
                            time = find_elm_with_attribute(li, lst_time)
                            description = find_elm_with_attribute(li, lst_description)
                            try:
                                center_element(driver.driver, li)
                                photo = find_elm_picture(
                                    li, Selector(css='img[src*=".jpg"]')
                                )
                                if photo:
                                    article.picture_filename = photo
                                    logger.info(
                                        f"Picture was found: {article.picture_filename}"
                                    )
                            except:
                                logger.info(
                                    "Information about picture in article was not found."
                                )
                                pass
                            logger.info("Information about article was found.")
                            article.title = title.text.strip()
                            article.description = description.text.strip()
                            article.date = datetime.strptime(
                                time.text.strip(), "%B %d, %Y"
                            )
                            logger.info(
                                f" Title: {article.title} -- Date: {article.date}"
                            )
                            sleep(0.4)
                            list_articles.append(article)
                        try:
                            button_next = find_element(
                                driver.driver,
                                Selector(css='button[aria-label*="Next stories"]'),
                            )
                            if button_next:
                                center_element(driver.driver, button_next)
                                click_elm(driver.driver, button_next)
                        except:
                            button_next = find_element(
                                driver.driver,
                                Selector(css='button[aria-label*="Disabled"]'),
                            )
                            if button_next:
                                more_results = False
            return list_articles
        except Exception as e:
            logger.critical(f"Some error occurred: {e.__cause__}.")
            return False


class ExcelOtherMethods:
    def __extract_filename_from_url(url):
        # Use regex to match a pattern that extracts the filename with extension
        match = re.search(r"[^/]+$", url)
        if match:
            filename_with_extension = match.group(0)
            return filename_with_extension
        else:
            return None

    def __contains_money(text):
        # Regular expression pattern to match monetary amounts
        pattern = r"\$[0-9,.]+|\b\d+\s*(?:dollars|USD)\b"

        # Use regular expression to search for matches in the text
        matches = re.findall(pattern, text)

        # If matches are found, return True; otherwise, return False
        return bool(matches)

    def __download_image(url):
        project_dir = str(os.getcwd())
        full_path = Path(project_dir, "devdata", "downloads")
        if os.path.isdir(full_path):
            full_path = os.path.join(
                full_path, ExcelOtherMethods.__extract_filename_from_url(url)
            )
            logger.info(f"Downloading image: {url}")
            urllib.request.urlretrieve(url, full_path)
            return full_path

    @staticmethod
    def prepare_articles(list_articles: list[Article], phrase: str) -> list[Article]:
        new_list_articles = []
        if list_articles:
            for article in list_articles:
                art = Article()
                art.title = article.title
                art.date = article.date
                art.title_count_phrase = article.title.lower().count(phrase.lower())
                art.description = ""
                art.description_count_phrase = 0
                art.find_money_title_description = ExcelOtherMethods.__contains_money(
                    article.title
                )
                # download image
                if len(article.picture_filename) > 0:
                    art.picture_filename = article.picture_filename
                    art.picture_local_path = ExcelOtherMethods.__download_image(
                        art.picture_filename
                    )
                new_list_articles.append(art)
                logger.info(f"Article created: {art}")
            return new_list_articles

    @staticmethod
    def export_excel(list_articles: list[Article]):
        project_dir = str(os.getcwd())
        full_path = Path(project_dir, "output")
        excel_file_path = os.path.join(full_path, "Articles.xlsx")
        wb = Workbook()
        ws = wb.active
        str_data = Article.articles_to_json(list_articles)
        data = json.loads(str_data)
        headers = list(data[0].keys()) if data else []
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)
        for row_index, row_data in enumerate(data, start=2):
            for col_index, header in enumerate(headers, start=1):
                ws.cell(row=row_index, column=col_index, value=row_data.get(header, ""))
        wb.save(excel_file_path)
        logger.info("Excel created")
        logger.info("Creating Output...")
        workitems.outputs.create(files=[excel_file_path])
